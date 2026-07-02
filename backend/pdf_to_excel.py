# pdf_to_excel.py

import argparse
import os
from copy import copy

import openpyxl
from img2table.document import PDF
from img2table.ocr import PaddleOCR

parser = argparse.ArgumentParser()
parser.add_argument("--input", type=str, default=None)
parser.add_argument("--engine", type=str, default="paddle", choices=["paddle", "easy"])
args, _ = parser.parse_known_args()

SAMPLES_DIR = "samples"
OUTPUT_DIR = "outputs"
# Daftar file default jika tidak ada --input (kosong = harus pakai --input)
TARGET_FILES = []


def get_output_paths(input_path: str):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = os.path.basename(input_path)
    base_name = os.path.splitext(filename)[0]
    final_path = os.path.join(OUTPUT_DIR, f"{base_name}.xlsx")
    temp_path = os.path.join(OUTPUT_DIR, f"temp_{base_name}.xlsx")
    return temp_path, final_path


def copy_sheet_content_no_merge(source_ws, target_ws, row_offset):
    for row in source_ws.iter_rows():
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            new_cell = target_ws.cell(
                row=cell.row + row_offset, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.alignment = copy(cell.alignment)
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)


def pdf_to_excel(pdf_path: str, engine: str = "paddle") -> str:
    print(f"\n[pdf_to_excel] Memproses: {pdf_path} (Engine: {engine})")
    temp_file, final_file = get_output_paths(pdf_path)

    if engine.lower() == "easy":
        from img2table.ocr import EasyOCR
        ocr = EasyOCR(lang=["en"])
    else:
        ocr = PaddleOCR(lang="en")

    try:
        doc = PDF(src=pdf_path)
        doc.to_xlsx(
            dest=temp_file,
            ocr=ocr,
            implicit_rows=False,
            borderless_tables=False,
            min_confidence=50,
        )
        print(f"  [1/2] OCR selesai. Temp: {temp_file}")
    except Exception as e:
        print(f"  [ERROR] OCR gagal: {e}")
        return None

    print(f"  [2/2] Meratakan sheet (tanpa merge)...")
    if not os.path.exists(temp_file):
        print("  [ERROR] File sementara tidak ditemukan.")
        return None

    wb_source = openpyxl.load_workbook(temp_file)
    wb_target = openpyxl.Workbook()
    ws_target = wb_target.active
    ws_target.title = "Data Flat"

    current_offset = 0
    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        copy_sheet_content_no_merge(ws_source, ws_target, current_offset)
        current_offset += ws_source.max_row + 3

    wb_target.save(final_file)
    print(f"  ✅ Tersimpan: {final_file}")

    wb_source.close()
    if os.path.exists(temp_file):
        os.remove(temp_file)

    return final_file


if __name__ == "__main__":
    # Jika ada --input dari pipeline, pakai itu. Kalau tidak, pakai TARGET_FILES
    if args.input:
        pdf_to_excel(args.input, engine=args.engine)
    elif TARGET_FILES:
        for filename in TARGET_FILES:
            pdf_path = os.path.join(SAMPLES_DIR, filename)
            if not os.path.exists(pdf_path):
                print(f"[SKIP] File tidak ditemukan: {pdf_path}")
                continue
            pdf_to_excel(pdf_path, engine=args.engine)
    else:
        print("[INFO] Tidak ada file yang diproses. Gunakan --input <path_pdf> atau isi TARGET_FILES.")
